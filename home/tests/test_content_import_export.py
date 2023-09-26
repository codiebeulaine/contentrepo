import csv
import json
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from functools import wraps
from io import BytesIO, StringIO
from pathlib import Path
from queue import Queue
from typing import Any

import pytest
from django.core import serializers  # type: ignore
from openpyxl import load_workbook
from wagtail.models import Locale, Page  # type: ignore
from wagtail.models.sites import Site  # type: ignore

from home.content_import_export import import_content, old_import_content
from home.models import ContentPage, ContentPageIndex, HomePage, SiteSettings

from .page_builder import MBlk, MBody, PageBuilder, VarMsg, VBlk, VBody, WABlk, WABody

ExpDict = dict[str, Any]
ExpPair = tuple[ExpDict, ExpDict]
ExpDicts = Iterable[ExpDict]
ExpDictsPair = tuple[ExpDicts, ExpDicts]


def filter_both(
    filter_func: Callable[[ExpDict], ExpDict]
) -> Callable[[ExpDict, ExpDict], ExpPair]:
    @wraps(filter_func)
    def ff(src: ExpDict, dst: ExpDict) -> ExpPair:
        return filter_func(src), filter_func(dst)

    return ff


@filter_both
def add_new_fields(entry: ExpDict) -> ExpDict:
    # FIXME: This should probably be in a separate test for importing old exports.
    return {"whatsapp_template_name": "", **entry}


def remove_translation_tag_from_tags(src: ExpDict, dst: ExpDict) -> ExpPair:
    # FIXME: Do we actually need translation_tag to be added to tags?
    if not src["translation_tag"]:
        return src, dst
    dtags = [tag for tag in dst["tags"].split(", ") if tag != src["translation_tag"]]
    return src, dst | {"tags": ", ".join(dtags)}


@filter_both
def ignore_certain_fields(entry: ExpDict) -> ExpDict:
    # FIXME: Do we need page.id to be imported? At the moment nothing in the
    #        import reads that.
    # FIXME: Implement import/export for doc_link, image_link, media_link.
    ignored_fields = {
        "page_id",
        "doc_link",
        "image_link",
        "media_link",
    }
    return {k: v for k, v in entry.items() if k not in ignored_fields}


@filter_both
def ignore_old_fields(entry: ExpDict) -> ExpDict:
    ignored_fields = {"next_prompt", "translation_tag"}
    return {k: v for k, v in entry.items() if k not in ignored_fields}


@filter_both
def strip_leading_whitespace(entry: ExpDict) -> ExpDict:
    # FIXME: Do we expect imported content to have leading spaces removed?
    bodies = {k: v.lstrip(" ") for k, v in entry.items() if k.endswith("_body")}
    return {**entry, **bodies}


EXPORT_FILTER_FUNCS = [
    add_new_fields,
    ignore_certain_fields,
    strip_leading_whitespace,
]

OLD_EXPORT_FILTER_FUNCS = [
    remove_translation_tag_from_tags,
    ignore_old_fields,
]


def filter_exports(srcs: ExpDicts, dsts: ExpDicts, importer: str) -> ExpDictsPair:
    fsrcs, fdsts = [], []
    for src, dst in zip(srcs, dsts, strict=True):
        for ff in EXPORT_FILTER_FUNCS:
            src, dst = ff(src, dst)
        if importer == "old":
            for ff in OLD_EXPORT_FILTER_FUNCS:
                src, dst = ff(src, dst)
        fsrcs.append(src)
        fdsts.append(dst)
    return fsrcs, fdsts


def csv2dicts(csv_bytes: bytes) -> ExpDicts:
    return list(csv.DictReader(StringIO(csv_bytes.decode())))


DbDict = dict[str, Any]
DbDicts = Iterable[DbDict]


def _models2dicts(model_instances: Any) -> DbDicts:
    return json.loads(serializers.serialize("json", model_instances))


def get_page_json() -> DbDicts:
    page_objs = Page.objects.type(ContentPage, ContentPageIndex).all()
    pages = {p["pk"]: p["fields"] for p in _models2dicts(page_objs)}
    content_pages = [
        *_models2dicts(ContentPage.objects.all()),
        *_models2dicts(ContentPageIndex.objects.all()),
    ]
    return [p | {"fields": {**pages[p["pk"]], **p["fields"]}} for p in content_pages]


def per_page(filter_func: Callable[[DbDict], DbDict]) -> Callable[[DbDicts], DbDicts]:
    @wraps(filter_func)
    def fp(pages: DbDicts) -> DbDicts:
        return [filter_func(page) for page in pages]

    return fp


@per_page
def bodies_to_dicts(page: DbDict) -> DbDict:
    fields = {
        k: json.loads(v) if k.endswith("body") else v for k, v in page["fields"].items()
    }
    return page | {"fields": fields}


def normalise_pks(pages: DbDicts) -> DbDicts:
    min_pk = min(p["pk"] for p in pages)
    return [p | {"pk": p["pk"] - min_pk} for p in pages]


def _update_field(
    pages: DbDicts, field_name: str, update_fn: Callable[[Any], Any]
) -> DbDicts:
    for p in pages:
        fields = p["fields"]
        yield p | {"fields": {**fields, field_name: update_fn(fields[field_name])}}


def normalise_revisions(pages: DbDicts) -> DbDicts:
    if "latest_revision" not in list(pages)[0]["fields"]:
        return pages
    min_latest = min(p["fields"]["latest_revision"] for p in pages)
    min_live = min(p["fields"]["live_revision"] for p in pages)
    pages = _update_field(pages, "latest_revision", lambda v: v - min_latest)
    pages = _update_field(pages, "live_revision", lambda v: v - min_live)
    return pages


def _remove_fields(pages: DbDicts, field_names: set[str]) -> DbDicts:
    for p in pages:
        fields = {k: v for k, v in p["fields"].items() if k not in field_names}
        yield p | {"fields": fields}


PAGE_TIMESTAMP_FIELDS = {
    "first_published_at",
    "last_published_at",
    "latest_revision_created_at",
}


def remove_timestamps(pages: DbDicts) -> DbDicts:
    return _remove_fields(pages, PAGE_TIMESTAMP_FIELDS)


def _normalise_varmsg_ids(page_id: str, var_list: list[dict[str, Any]]) -> None:
    for i, varmsg in enumerate(var_list):
        assert "id" in varmsg
        varmsg["id"] = f"{page_id}:var:{i}"
        for ir, rest in enumerate(varmsg["value"]["variation_restrictions"]):
            rest["id"] = f"{page_id}:var:{i}:var:{ir}"


def _normalise_body_field_ids(
    page: DbDict, body_name: str, body_list: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    for i, body in enumerate(body_list):
        assert "id" in body
        body["id"] = f"fake:{page['pk']}:{body_name}:{i}"
        if "variation_messages" in body["value"]:
            _normalise_varmsg_ids(body["id"], body["value"]["variation_messages"])
    return body_list


@per_page
def normalise_body_ids(page: DbDict) -> DbDict:
    # FIXME: Does it matter if these change?
    fields = {
        k: _normalise_body_field_ids(page, k, v) if k.endswith("body") else v
        for k, v in page["fields"].items()
    }
    return page | {"fields": fields}


def remove_translation_key(pages: DbDicts) -> DbDicts:
    # For old importer.
    return _remove_fields(pages, {"translation_key"})


def remove_revisions(pages: DbDicts) -> DbDicts:
    # For old importer. Sometimes (maybe for the ContentPages imported after
    # the first language?) we get higher revision numbers. Let's just strip
    # them all and be done with it.
    return _remove_fields(pages, {"latest_revision", "live_revision"})


@per_page
def null_to_emptystr(page: DbDict) -> DbDict:
    # FIXME: Confirm that there's no meaningful difference here, potentially
    #        make these fields non-nullable.
    fields = {**page["fields"]}
    for k in ["subtitle", "whatsapp_title", "messenger_title", "viber_title"]:
        if k in fields and fields[k] is None:
            fields[k] = ""
    if "whatsapp_body" in fields:
        for body in fields["whatsapp_body"]:
            if "next_prompt" in body["value"] and not body["value"]["next_prompt"]:
                body["value"]["next_prompt"] = ""
    return page | {"fields": fields}


def _add_fields(body: dict[str, Any], extra_fields: dict[str, Any]) -> None:
    body["value"] = extra_fields | body["value"]


@per_page
def add_body_fields(page: DbDict) -> DbDict:
    if "whatsapp_body" in page["fields"]:
        for body in page["fields"]["whatsapp_body"]:
            _add_fields(
                body,
                {
                    "document": None,
                    "image": None,
                    "media": None,
                    "next_prompt": "",
                    "variation_messages": [],
                },
            )
    if "messenger_body" in page["fields"]:
        for body in page["fields"]["messenger_body"]:
            _add_fields(body, {"image": None})
    if "viber_body" in page["fields"]:
        for body in page["fields"]["viber_body"]:
            _add_fields(body, {"image": None})
    return page


@per_page
def remove_next_prompt(page: DbDict) -> DbDict:
    if "whatsapp_body" in page["fields"]:
        for body in page["fields"]["whatsapp_body"]:
            body["value"].pop("next_prompt", None)
    return page


@per_page
def enable_web(page: DbDict) -> DbDict:
    page["fields"]["enable_web"] = True
    return page


PAGE_FILTER_FUNCS = [
    normalise_pks,
    normalise_revisions,
    remove_timestamps,
    normalise_body_ids,
    null_to_emptystr,
]

OLD_PAGE_FILTER_FUNCS = [
    remove_translation_key,
    remove_revisions,
    add_body_fields,
    remove_next_prompt,
    enable_web,
]


PFOption = tuple[str, list[str]]
PFOptions = list[PFOption]


def set_profile_field_options(profile_field_options: PFOptions) -> None:
    site = Site.objects.get(is_default_site=True)
    site_settings = SiteSettings.for_site(site)
    site_settings.profile_field_options.extend(profile_field_options)
    site_settings.save()


@dataclass
class ImportExportFixture:
    admin_client: Any
    importer: str
    format: str

    @property
    def _import_content(self) -> Callable[..., None]:
        return {
            "new": import_content,
            "old": old_import_content,
        }[self.importer]

    @property
    def _filter_export(self) -> Callable[..., bytes]:
        return {
            "csv": self._filter_export_CSV,
            "xlsx": self._filter_export_XLSX,
        }[self.format]

    def _filter_export_row(self, row: ExpDict, locale: str | None) -> bool:
        """
        Determine whether to keep a given export row.
        """
        if locale:
            if row["locale"] not in [None, "", locale]:
                return False
        return True

    def _filter_export_CSV(self, content: bytes, locale: str | None) -> bytes:
        reader = csv.DictReader(StringIO(content.decode()))
        assert reader.fieldnames is not None
        out_content = StringIO()
        writer = csv.DictWriter(out_content, fieldnames=reader.fieldnames)
        writer.writeheader()
        for row in reader:
            if self._filter_export_row(row, locale=locale):
                writer.writerow(row)
        return out_content.getvalue().encode()

    def _filter_export_XLSX(self, content: bytes, locale: str | None) -> bytes:
        workbook = load_workbook(BytesIO(content))
        worksheet = workbook.worksheets[0]
        header = next(worksheet.iter_rows(max_row=1, values_only=True))

        rows_to_remove: list[int] = []
        for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
            r = {k: v for k, v in zip(header, row, strict=True) if isinstance(k, str)}
            if not self._filter_export_row(r, locale=locale):
                rows_to_remove.append(i + 2)
        for row_num in reversed(rows_to_remove):
            worksheet.delete_rows(row_num)

        out_content = BytesIO()
        workbook.save(out_content)
        return out_content.getvalue()

    def export_content(self, locale: str | None = None) -> bytes:
        """
        Export all (or filtered) content in the configured format.

        FIXME:
         * If we filter the export by locale, we only get ContentPage entries
           for the given language, but we still get ContentPageIndex rows for
           all languages.
        """
        url = f"/admin/home/contentpage/?export={self.format}"
        if locale:
            loc = Locale.objects.get(language_code=locale)
            locale = str(loc)
            url = f"{url}&locale__id__exact={loc.id}"
        content = self.admin_client.get(url).content
        # Hopefully we can get rid of this at some point.
        if locale:
            content = self._filter_export(content, locale=locale)
        return content

    def import_content(self, content_bytes: bytes, **kw: Any) -> None:
        """
        Import given content in the configured format with the configured importer.
        """
        self._import_content(BytesIO(content_bytes), self.format.upper(), Queue(), **kw)

    def import_file(
        self, path_str: str, path_base: str = "home/tests", **kw: Any
    ) -> bytes:
        """
        Import given content file in the configured format with the configured importer.
        """
        content = (Path(path_base) / path_str).read_bytes()
        self.import_content(content, **kw)
        return content

    def export_reimport(self) -> None:
        """
        Export all content, then immediately reimport it.
        """
        self.import_content(self.export_content())

    def get_page_json(self) -> DbDicts:
        """
        Serialize all ContentPage and ContentPageIndex instances and normalize
        things that vary across import/export.
        """
        pages = bodies_to_dicts(get_page_json())
        if self.importer == "old":
            for ff in OLD_PAGE_FILTER_FUNCS:
                pages = ff(pages)
        for ff in PAGE_FILTER_FUNCS:
            pages = ff(pages)
        return sorted(pages, key=lambda p: p["pk"])

    def csvs2dicts(self, src_bytes: bytes, dst_bytes: bytes) -> ExpDictsPair:
        src = csv2dicts(src_bytes)
        dst = csv2dicts(dst_bytes)
        return filter_exports(src, dst, self.importer)


@pytest.fixture(params=["old", "new"])
def csv_impexp(request: Any, admin_client: Any) -> ImportExportFixture:
    return ImportExportFixture(admin_client, request.param, "csv")


@pytest.mark.django_db
class TestImportExportRoundtrip:
    """
    Test importing and reexporting content produces an export that is
    equilavent to the original imported content.

    NOTE: This is not a Django (or even unittest) TestCase. It's just a
        container for related tests.
    """

    def test_roundtrip_csv_simple(self, csv_impexp: ImportExportFixture) -> None:
        """
        Importing a simple CSV file and then exporting it produces a duplicate
        of the original file.

        (This uses content2.csv from test_api.py.)

        FIXME:
         * This should probably be in a separate test for importing old exports.
         * Do we actually need translation_tag to be added to tags?
         * Do we need page.id to be imported? At the moment nothing in the
           import reads that.
         * Do we expect imported content to have leading spaces removed?
         * Should we set enable_web and friends based on body, title, or an
           enable field that we'll need to add to the export?
        """
        csv_bytes = csv_impexp.import_file("content2.csv")
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_roundtrip_csv_less_simple(self, csv_impexp: ImportExportFixture) -> None:
        """
        Importing a less simple CSV file and then exporting it produces a
        duplicate of the original file.

        (This uses exported_content_20230911-variations-linked-page.csv.)

        FIXME:
         * Implement import/export for doc_link, image_link, media_link.
        """
        set_profile_field_options([("gender", ["male", "female", "empty"])])
        csv_bytes = csv_impexp.import_file(
            "exported_content_20230911-variations-linked-page.csv"
        )
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_roundtrip_csv_translations(self, csv_impexp: ImportExportFixture) -> None:
        """
        Importing a CSV file containing translations and then exporting it
        produces a duplicate of the original file.

        (This uses exported_content_20230906-translations.csv and the two
        language-specific subsets thereof.)

        FIXME:
         * We shouldn't need to import different languages separately.
         * Slugs need to either be unique per site/deployment or the importer
           needs to handle them only being unique per locale.
        """

        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        set_profile_field_options([("gender", ["male", "female", "empty"])])
        csv_impexp.import_file("translations-en.csv")
        csv_impexp.import_file("translations-pt.csv", locale="pt", purge=False)
        csv_bytes = Path(
            "home/tests/exported_content_20230906-translations.csv"
        ).read_bytes()
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src

    def test_import_error(self, csv_impexp: ImportExportFixture) -> None:
        """
        Importing an invalid CSV file leaves the db as-is.

        (This uses content2.csv from test_api.py and broken.csv.)
        """
        # Start with some existing content.
        csv_bytes = csv_impexp.import_file("content2.csv")

        # This CSV doesn't have any of the fields we expect.
        with pytest.raises((KeyError, TypeError)):
            csv_impexp.import_file("broken.csv")

        # The export should match the existing content.
        content = csv_impexp.export_content()
        src, dst = csv_impexp.csvs2dicts(csv_bytes, content)
        assert dst == src


# "old-xlsx" has at least three bugs, so we don't bother testing it.
@pytest.fixture(params=["old-csv", "new-csv", "new-xlsx"])
def impexp(request: Any, admin_client: Any) -> ImportExportFixture:
    importer, format = request.param.split("-")
    return ImportExportFixture(admin_client, importer, format)


@pytest.mark.django_db
class TestExportImportRoundtrip:
    """
    Test that the db state after exporting and reimporting content is
    equilavent to what it was before.

    NOTE: This is not a Django (or even unittest) TestCase. It's just a
        container for related tests.
    """

    def test_roundtrip_simple(self, impexp: ImportExportFixture) -> None:
        """
        Exporting and then importing leaves the db in the same state it was
        before, except for page_ids, timestamps, and body item ids.

        FIXME:
         * Determine whether we need to maintain StreamField block ids. (I
           think we don't.)
         * Confirm that there's no meaningful difference between null and ""
           for the nullable fields that the importer sets to "", potentially
           make these fields non-nullable.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert* WA")]),
                MBody("HealthAlert menu", [MBlk("Welcome to HealthAlert M")]),
            ],
        )
        _health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[
                WABody("health info", [WABlk("*Health information* WA")]),
                MBody("health info", [MBlk("*Health information* M")]),
            ],
        )
        _self_help = PageBuilder.build_cp(
            parent=ha_menu,
            slug="self-help",
            title="self-help",
            bodies=[
                WABody("self-help", [WABlk("*Self-help programs* WA")]),
                MBody("self-help", [MBlk("*Self-help programs* M")]),
                VBody("self-help", [VBlk("*Self-help programs* V")]),
            ],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_roundtrip_variations(self, impexp: ImportExportFixture) -> None:
        """
        ContentPages with variation messages (and next prompts) are preserved
        across export/import.
        """
        set_profile_field_options([("gender", ["male", "female", "empty"])])

        home_page = HomePage.objects.first()
        imp_exp = PageBuilder.build_cpi(home_page, "import-export", "Import Export")

        cp_imp_exp_wablks = [
            WABlk(
                "Message 1",
                next_prompt="Next message",
                variation_messages=[VarMsg("Var'n for Gender Male", gender="male")],
            ),
            WABlk(
                "Message2, variable placeholders as well {{0}}",
                next_prompt="Next message",
                variation_messages=[VarMsg("Var'n for Rather not say", gender="empty")],
            ),
            WABlk("Message 3 with no variation", next_prompt="end"),
        ]
        _cp_imp_exp = PageBuilder.build_cp(
            parent=imp_exp,
            slug="cp-import-export",
            title="CP-Import/export",
            bodies=[WABody("WA import export data", cp_imp_exp_wablks)],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_roundtrip_tags(self, impexp: ImportExportFixture) -> None:
        """
        ContentPages with tags are preserved across export/import.
        """
        home_page = HomePage.objects.first()
        main_menu = PageBuilder.build_cpi(home_page, "main-menu", "Main Menu")
        ha_menu = PageBuilder.build_cp(
            parent=main_menu,
            slug="ha-menu",
            title="HealthAlert menu",
            bodies=[
                WABody("HealthAlert menu", [WABlk("*Welcome to HealthAlert* WA")]),
                MBody("HealthAlert menu", [MBlk("Welcome to HealthAlert M")]),
            ],
            tags=["tag1", "tag2"],
        )
        _health_info = PageBuilder.build_cp(
            parent=ha_menu,
            slug="health-info",
            title="health info",
            bodies=[MBody("health info", [MBlk("*Health information* M")])],
            tags=["tag2", "tag3"],
        )
        _self_help = PageBuilder.build_cp(
            parent=ha_menu,
            slug="self-help",
            title="self-help",
            bodies=[WABody("self-help", [WABlk("*Self-help programs* WA")])],
            tags=["tag4"],
        )

        orig = impexp.get_page_json()
        impexp.export_reimport()
        imported = impexp.get_page_json()
        assert imported == orig

    def test_roundtrip_translations(self, impexp: ImportExportFixture) -> None:
        """
        ContentPages in multiple languages are preserved across export/import.

        FIXME:
         * We shouldn't need to import different languages separately.
        """
        # Create a new homepage for Portuguese.
        pt, _created = Locale.objects.get_or_create(language_code="pt")
        HomePage.add_root(locale=pt, title="Home (pt)", slug="home-pt")

        # English pages
        home_en = HomePage.objects.get(locale__language_code="en")
        _app_rem = PageBuilder.build_cpi(
            parent=home_en, slug="appointment-reminders", title="Appointment reminders"
        )
        _sbm = PageBuilder.build_cpi(
            parent=home_en, slug="stage-based-messages", title="Stage-based messages"
        )
        _him = PageBuilder.build_cpi(
            parent=home_en, slug="health-info-messages", title="Health info messages"
        )
        _wtt = PageBuilder.build_cpi(
            parent=home_en,
            slug="whatsapp-template-testing",
            title="whatsapp template testing",
        )
        imp_exp = PageBuilder.build_cpi(
            parent=home_en, slug="import-export", title="Import Export"
        )
        non_templ_wablks = [
            WABlk("this is a non template message"),
            WABlk("this message has a doc"),
            WABlk("this message comes with audio"),
        ]
        non_tmpl = PageBuilder.build_cp(
            parent=imp_exp,
            slug="non-template",
            title="Non template messages",
            bodies=[WABody("non template OCS", non_templ_wablks)],
        )

        # Portuguese pages
        home_pt = HomePage.objects.get(locale__language_code="pt")
        imp_exp_pt = PageBuilder.build_cpi(
            parent=home_pt,
            slug="import-export-pt",
            title="Import Export (pt)",
            translated_from=imp_exp,
        )
        non_templ_wablks_pt = [
            WABlk("this is a non template message (pt)"),
            WABlk("this message has a doc (pt)"),
            WABlk("this message comes with audio (pt)"),
        ]
        non_tmpl_pt = PageBuilder.build_cp(
            parent=imp_exp_pt,
            slug="non-template-pt",
            title="Non template messages",
            bodies=[WABody("non template OCS", non_templ_wablks_pt)],
            translated_from=non_tmpl,
        )

        assert imp_exp.translation_key == imp_exp_pt.translation_key
        assert non_tmpl.translation_key == non_tmpl_pt.translation_key

        orig = impexp.get_page_json()
        content_en = impexp.export_content(locale="en")
        content_pt = impexp.export_content(locale="pt")

        impexp.import_content(content_en, locale="en")
        impexp.import_content(content_pt, locale="pt", purge=False)
        imported = impexp.get_page_json()
        assert imported == orig
