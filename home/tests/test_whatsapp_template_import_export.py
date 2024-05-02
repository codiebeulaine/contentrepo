import csv
import json
import re
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from functools import wraps
from io import BytesIO, StringIO
from pathlib import Path
from queue import Queue
from typing import Any, TypeVar

import pytest
from django.core import serializers  # type: ignore
from openpyxl import load_workbook
from pytest_django.fixtures import SettingsWrapper
from wagtail.models import Locale  # type: ignore
from wagtail.snippets.models import register_snippet  # type: ignore

from home.import_whatsapp_templates import ImportWhatsAppTemplateException
from home.models import (
    WhatsAppTemplate,
)
from home.wagtail_hooks import WhatsAppTemplateAdmin
from home.whatsapp_template_import_export import import_whatsapptemplate

TTemplate = TypeVar("TTemplate", bound=WhatsAppTemplate)

IMP_EXP_DATA_BASE = Path("home/tests/import-export-data")

ExpDict = dict[str, Any]
ExpPair = tuple[ExpDict, ExpDict]
ExpDicts = Iterable[ExpDict]
ExpDictsPair = tuple[ExpDicts, ExpDicts]


def filter_both(
    filter_func: Callable[[ExpDict], ExpDict]
) -> Callable[[ExpDict, ExpDict], ExpPair]:
    @wraps(filter_func)
    def ff(src: ExpDict, dst: ExpDict) -> ExpPair:
        return filter_func(src), filter_func(dst)

    return ff


@filter_both
def strip_leading_whitespace(entry: ExpDict) -> ExpDict:
    # FIXME: Do we expect imported content to have leading spaces removed?
    bodies = {k: v.lstrip(" ") for k, v in entry.items() if k.endswith("_body")}
    return {**entry, **bodies}


EXPORT_FILTER_FUNCS = [
    strip_leading_whitespace,
]


def filter_exports(srcs: ExpDicts, dsts: ExpDicts) -> ExpDictsPair:
    fsrcs, fdsts = [], []
    for src, dst in zip(srcs, dsts, strict=True):
        for ff in EXPORT_FILTER_FUNCS:
            src, dst = ff(src, dst)
        fsrcs.append(src)
        fdsts.append(dst)
    return fsrcs, fdsts


def csv2dicts(csv_bytes: bytes) -> ExpDicts:
    return list(csv.DictReader(StringIO(csv_bytes.decode())))


WEB_PARA_RE = re.compile(r'^<div class="block-paragraph">(.*)</div>$')

DbDict = dict[str, Any]
DbDicts = Iterable[DbDict]


def _models2dicts(model_instances: Any) -> DbDicts:
    return json.loads(serializers.serialize("json", model_instances))


def get_whatsapp_template_json() -> DbDicts:
    templates = [*_models2dicts(WhatsAppTemplate.objects.all())]
    return templates


def _is_json_field(field_name: str) -> bool:
    return field_name in {"example_values"}


def per_template(
    filter_func: Callable[[DbDict], DbDict]
) -> Callable[[DbDicts], DbDicts]:
    @wraps(filter_func)
    def fp(templates: DbDicts) -> DbDicts:
        return [filter_func(template) for template in templates]

    return fp


@per_template
def decode_json_fields(template: DbDict) -> DbDict:

    fields = {
        k: json.loads(v) if _is_json_field(k) else v
        for k, v in template["fields"].items()
    }
    return template | {"fields": fields}


def _update_field(
    templates: DbDicts, field_name: str, update_fn: Callable[[Any], Any]
) -> DbDicts:
    for t in templates:
        fields = t["fields"]
        yield t | {"fields": {**fields, field_name: update_fn(fields[field_name])}}


def normalise_pks(templates: DbDicts) -> DbDicts:
    min_pk = min(t["pk"] for t in templates)
    return [_normalise_pks(template, min_pk) for template in templates]


def _normalise_pks(template: DbDict, min_pk: str) -> DbDict:
    return template | {"pk": template["pk"] - min_pk}


@per_template
def remove_example_value_ids(template: DbDict) -> DbDict:
    if "example_values" in template["fields"]:
        example_values = template["fields"]["example_values"]
        for example_value in example_values:
            example_value.pop("id", None)

    return template


WHATSAPP_TEMPLATE_FILTER_FUNCS = [remove_example_value_ids, normalise_pks]


@dataclass
class ImportExport:
    admin_client: Any
    format: str

    @property
    def _filter_export(self) -> Callable[..., bytes]:
        return {
            "csv": self._filter_export_CSV,
            "xlsx": self._filter_export_XLSX,
        }[self.format]

    def _filter_export_row(self, row: ExpDict, locale: str | None) -> bool:
        """
        Determine whether to keep a given export row.
        """
        if locale:
            if row["locale"] not in [None, "", locale]:
                return False
        return True

    def _filter_export_CSV(self, content: bytes, locale: str | None) -> bytes:
        reader = csv.DictReader(StringIO(content.decode()))
        assert reader.fieldnames is not None
        out_content = StringIO()
        writer = csv.DictWriter(out_content, fieldnames=reader.fieldnames)
        writer.writeheader()
        for row in reader:
            if self._filter_export_row(row, locale=locale):
                writer.writerow(row)
        return out_content.getvalue().encode()

    def _filter_export_XLSX(self, content: bytes, locale: str | None) -> bytes:
        workbook = load_workbook(BytesIO(content))
        worksheet = workbook.worksheets[0]
        header = next(worksheet.iter_rows(max_row=1, values_only=True))

        rows_to_remove: list[int] = []
        for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
            r = {k: v for k, v in zip(header, row, strict=True) if isinstance(k, str)}
            if not self._filter_export_row(r, locale=locale):
                rows_to_remove.append(i + 2)
        for row_num in reversed(rows_to_remove):
            worksheet.delete_rows(row_num)

        out_content = BytesIO()
        workbook.save(out_content)
        return out_content.getvalue()

    def export_whatsapp_template(self) -> bytes:
        """
        Export all (or filtered) content in the configured format.
        """
        url = f"/admin/snippets/home/whatsapptemplate/?export={self.format}"
        content = self.admin_client.get(url).content
        # Hopefully we can get rid of this at some point.
        print(f"Content = {content}")
        if self.format == "csv":
            print("-v-CONTENT-v-")
            print(content.decode())
            print("-^-CONTENT-^-")
        return content

    def import_whatsapp_template(self, content_bytes: bytes, **kw: Any) -> None:
        """
        Import given content in the configured format with the configured importer.
        """
        import_whatsapptemplate(
            BytesIO(content_bytes), self.format.upper(), Queue(), **kw
        )

    def read_bytes(self, path_str: str, path_base: Path = IMP_EXP_DATA_BASE) -> bytes:
        return (path_base / path_str).read_bytes()

    def import_file(
        self, path_str: str, path_base: Path = IMP_EXP_DATA_BASE, **kw: Any
    ) -> bytes:
        """
        Import given content file in the configured format with the configured importer.
        """
        content = self.read_bytes(path_str, path_base)
        self.import_whatsapp_template(content, **kw)
        return content

    def export_reimport(self) -> None:
        """
        Export all content, then immediately reimport it.
        """
        self.import_whatsapp_template(self.export_whatsapp_template())

    def csvs2dicts(self, src_bytes: bytes, dst_bytes: bytes) -> ExpDictsPair:
        src = csv2dicts(src_bytes)
        dst = csv2dicts(dst_bytes)
        return filter_exports(src, dst)

    def get_whatsapp_template_json(self) -> DbDicts:
        """
        Serialize all Whatsapp Template instances and normalize things that vary across
        import/export.
        """
        templates = decode_json_fields(get_whatsapp_template_json())
        for ff in WHATSAPP_TEMPLATE_FILTER_FUNCS:
            templates = ff(templates)
        return sorted(templates, key=lambda p: p["pk"])


@pytest.fixture(params=["csv", "xlsx"])
def impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, request.param)


@pytest.fixture()
def tmp_media_path(tmp_path: Path, settings: SettingsWrapper) -> None:
    settings.MEDIA_ROOT = tmp_path


@pytest.fixture()
def csv_impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, "csv")


@pytest.fixture()
def xlsx_impexp(request: Any, admin_client: Any) -> ImportExport:
    return ImportExport(admin_client, "xlsx")


# Needs this here while standalone templates are hidden behind feature flag
register_snippet(WhatsAppTemplateAdmin)


@pytest.mark.django_db()
class TestImportExportRoundtrip:
    """
    Test importing and reexporting content produces an export that is
    equilavent to the original imported content.

    NOTE: This is not a Django (or even unittest) TestCase. It's just a
        container for related tests.
    """

    def test_simple(self, csv_impexp: ImportExport) -> None:
        """
        Importing a simple CSV file with one whatsapp template and
        one question and export it

        (This uses whatsapp_template_simple.csv.)

        """

        csv_bytes = csv_impexp.import_file("whatsapp-template-simple.csv")
        content = csv_impexp.export_whatsapp_template()
        csv, export = csv_impexp.csvs2dicts(csv_bytes, content)
        assert export == csv

    def test_less_simple_with_quick_replies(self, csv_impexp: ImportExport) -> None:
        """
        Importing a simple CSV file with one whatsapp templates including quick replies


        (This uses whatsapp-template-less-simple.csv.)

        """
        csv_bytes = csv_impexp.import_file("whatsapp-template-less-simple.csv")
        content = csv_impexp.export_whatsapp_template()
        csv, export = csv_impexp.csvs2dicts(csv_bytes, content)
        assert export == csv

    def test_multiple_whatsapp_templates(self, csv_impexp: ImportExport) -> None:
        """
        Importing a simple CSV file with more than one whatsapp template and export it

        (This uses whatsapp-template-multiple.csv)

        """
        csv_bytes = csv_impexp.import_file("whatsapp-template-multiple.csv")
        content = csv_impexp.export_whatsapp_template()
        csv, export = csv_impexp.csvs2dicts(csv_bytes, content)
        assert export == csv

    def test_single_whatsapp_template(self, impexp: ImportExport) -> None:
        """
        Exporting then reimporting leaves the database in the same state we started with
        """
        WhatsAppTemplate.objects.create(
            name="wa_title",
            message="Test WhatsApp Message with two placeholders {{1}} and {{2}}",
            category="UTILITY",
            locale=Locale.objects.get(language_code="en"),
            example_values=[
                ("example_values", "Ev1"),
                ("example_values", "Ev2"),
            ],
            submission_name="testname",
            submission_status="NOT YET SUBMITTED",
            submission_result="test result",
        )

        orig = impexp.get_whatsapp_template_json()
        impexp.export_reimport()
        imported = impexp.get_whatsapp_template_json()
        assert imported == orig


@pytest.mark.django_db()
class TestImportExport:
    """
    Test import and export scenarios that aren't specifically round
    trips.
    """

    def test_import_error(elf, csv_impexp: ImportExport) -> None:
        """
        Importing an invalid CSV file leaves the db as-is.
        """
        csv_bytes = csv_impexp.import_file("whatsapp-template-simple.csv")
        with pytest.raises(ImportWhatsAppTemplateException) as e:
            csv_impexp.import_file("whatsapp-template-broken.csv")
        assert e.value.message == "Language code not found: "
        content = csv_impexp.export_whatsapp_template()
        csv, export = csv_impexp.csvs2dicts(csv_bytes, content)
        assert export == csv

    def test_invalid_locale_code(self, csv_impexp: ImportExport) -> None:
        """
        Importing whatsapp templates with invalid locale code should raise an error that results
        in an error message that gets sent back to the user
        """
        with pytest.raises(ImportWhatsAppTemplateException) as e:
            csv_impexp.import_file("invalid-whatsapp-template-locale-name.csv")

        assert e.value.row_num == 2
        assert e.value.message == "Language code not found: fakecode"
