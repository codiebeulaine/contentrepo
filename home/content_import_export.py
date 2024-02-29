import copy
import csv
import io
from dataclasses import dataclass
from io import BytesIO
from json import dumps
from logging import getLogger
from math import ceil
from typing import List, Tuple, Union

from django.db import transaction
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Border, Color, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from wagtail import blocks
from wagtail.documents.models import Document
from wagtail.images.models import Image
from wagtail.models import Locale
from wagtail.query import PageQuerySet
from wagtailmedia.models import Media

from home.models import (  # isort:skip
    ContentPage,
    ContentPageIndex,
    HomePage,
    OrderedContentSet,
)


EXPORT_FIELDNAMES = [
    "page_id",
    "slug",
    "parent",
    "web_title",
    "web_subtitle",
    "web_body",
    "whatsapp_title",
    "whatsapp_body",
    "whatsapp_template_name",
    "whatsapp_template_category",
    "example_values",
    "variation_title",
    "variation_body",
    "list_items",
    "sms_title",
    "sms_body",
    "ussd_title",
    "ussd_body",
    "messenger_title",
    "messenger_body",
    "viber_title",
    "viber_body",
    "translation_tag",
    "tags",
    "quick_replies",
    "triggers",
    "locale",
    "next_prompt",
    "buttons",
    "image_link",
    "doc_link",
    "media_link",
    "related_pages",
]

logger = getLogger(__name__)


@transaction.atomic
def import_content(file, filetype, progress_queue, purge=True, locale=None) -> None:
    from .import_content_pages import ContentImporter

    importer = ContentImporter(file.read(), filetype, progress_queue, purge, locale)
    importer.perform_import()


def export_xlsx_content(queryset: PageQuerySet, response: HttpResponse) -> None:
    from .export_content_pages import ContentExporter, ExportWriter

    exporter = ContentExporter(queryset)
    export_rows = exporter.perform_export()
    ExportWriter(export_rows).write_xlsx(response)


def export_csv_content(queryset: PageQuerySet, response: HttpResponse) -> None:
    from .export_content_pages import ContentExporter, ExportWriter

    exporter = ContentExporter(queryset)
    export_rows = exporter.perform_export()
    ExportWriter(export_rows).write_csv(response)


def import_ordered_sets(file, filetype, progress_queue, purge=False):
    def create_ordered_set_from_row(row):
        set_name = row["Name"]

        ordered_set = OrderedContentSet.objects.filter(name=set_name).first()
        if not ordered_set:
            ordered_set = OrderedContentSet(name=set_name)

        ordered_set.profile_fields = []
        for field in [f.strip() for f in row["Profile Fields"].split(",")]:
            if not field or field == "-":
                continue
            [field_name, field_value] = field.split(":")
            ordered_set.profile_fields.append((field_name, field_value))

        ordered_set.pages = []
        for page_slug in [p.strip() for p in row["Page Slugs"].split(",")]:
            if not page_slug or page_slug == "-":
                continue
            page = ContentPage.objects.filter(slug=page_slug).first()
            if page:
                ordered_set.pages.append(
                    (
                        "pages",
                        {
                            "contentpage": page,
                            "time": row["Time"] or "",
                            "unit": row["Unit"] or "",
                            "before_or_after": row["Before Or After"] or "",
                            "contact_field": row["Contact Field"] or "",
                        },
                    )
                )
            else:
                logger.warning(f"Content page not found for slug '{page_slug}'")

        ordered_set.save()
        return ordered_set

    file = file.read()
    lines = []
    if filetype == "XLSX":
        wb = load_workbook(filename=BytesIO(file))
        ws = wb.worksheets[0]
        ws.delete_rows(1)
        for row in ws.iter_rows(values_only=True):
            row_dict = {
                "Name": row[0],
                "Profile Fields": row[1],
                "Page Slugs": row[2],
                "Time": row[3],
                "Unit": row[4],
                "Before Or After": row[5],
                "Contact Field": row[6],
            }
            lines.append(row_dict)
    else:
        if isinstance(file, bytes):
            try:
                file = file.decode("utf-8")
            except UnicodeDecodeError:
                file = file.decode("latin-1")

        reader = csv.DictReader(io.StringIO(file))
        for dictionary in reader:
            lines.append(dictionary)

    # 10% progress for loading file
    progress_queue.put_nowait(10)

    for index, row in enumerate(lines):
        os = create_ordered_set_from_row(row)
        if not os:
            print(f"Ordered Content Set not created for row {index + 1}")
        # 10-100% for loading ordered content sets
        progress_queue.put_nowait(10 + index * 90 / len(lines))
